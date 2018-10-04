#使用各种模块练习对Excel的操作
#xlrd xlwt openpyxl
#C:\Users\mengmengli>pip install xlrd
#C:\Users\mengmengli>pip install xlwt
#C:\Users\mengmengli>pip install openpyxl

#2007版以前的Excel（xls结尾的），需要使用xlrd读，xlwt写
#import xlwt #Microsoft Excel versions 95 to 2003
#import xlrd #Microsoft Excel versions 95 to 2003
#2007版以后的Excel（xlsx结尾的），需要使用openpyxl来读写
import openpyxl

def writeExcel(path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "测试表A"

    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]

    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))

    wb.save(path)
    print("写入数据成功！")

def readExcel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name("测试表A")
    #sheet1 = wb.worksheets
    print(sheet)

    for row in sheet:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

if __name__=="__main__":
    file = 'E:/demo/Python/Excel/图书信息.xlsx'

    writeExcel(file)
    readExcel(file)