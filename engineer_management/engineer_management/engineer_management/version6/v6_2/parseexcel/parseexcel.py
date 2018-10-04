'''
from openpyxl import load_workbook
from openpyxl import Workbook
class ParseExcel:
    def __init__(self,path,sheetname):
        self.wb=load_workbook(path)
        self.sheet=self.wb.get_sheet_by_name(sheetname)
        self.maxrow=self.sheet.max_row

    # 写入数据
    def write_data(self):
        #创建一个新的工作表
        wb=Workbook()
        print(wb.get_sheet_names)
        wb.shee

    #读取数据
    def read_data(self):
        datalist=list()
        for line in self.sheet.rows[1:]:

'''



