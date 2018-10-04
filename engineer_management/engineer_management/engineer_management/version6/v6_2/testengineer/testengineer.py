import copy
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import worksheet
class TestEngineer():
    #构造函数，初始化数据
    def __init__(self):
        #存储输入的每个工程师信息
        self.engineer_information = dict()
        # 工程师信息初始化
        self.init_data()
        #版本5_1为了存储多人信息使用的list1、list2、h
        #暂存新添加的每个工程师信息，准备插入到list2中
        self.list1 = list()
        #存储所有工程师信息，每个元素为一个字典
        self.list2 = list()
        # 记录总共人数
        self.h=0
        # 版本5_2实现，启动时直接导入文件中的数据
        self.import_data()

    # 功能1：版本5_1,可以输入多个工程师信息
    def input_information(self):
        print("-------------------------------")
        an2="y"
        #通过循环输入多人信息
        while an2 in("y","Y"):
            self.h += 1
            print("请输入第%d个软件测试工程师的信息:"%self.h)
            self.engineer_information["engineer_id"] = input("请输入编号:")
            self.engineer_information["engineer_name"] = input("请输入姓名：")
            self.engineer_information["engineer_sex"] = input("请输入性别：")
            self.engineer_information["engineer_work_year"] = input("请输入工龄：")
            self.engineer_information["engineer_base_salary"] = input("请输入基本工资：")
            self.engineer_information["engineer_performance"] = input("请输入绩效：")
            self.engineer_information["engineer_workday"] = input("请输入月有效工作天数：")
            self.engineer_information["engineer_InsuranceFee"] = input("请输入月保险费：")
            self.engineer_information["engineer_foodfee"]=0.00
            self.engineer_information["engineer_foodfee"] = 0.00
            self.engineer_information["total_salary"] = 0.00
            # 版本5_1实现，在list2中存储的多人信息，每个信息是一个字典
            #把字典深复制到列表1中（list1转换为字典）
            self.list1=copy.deepcopy(self.engineer_information)
            # 将列表1插入到列表2指定位置
            self.list2.insert(self.h, self.list1)
            an2=input("是否继续输入？(y/n)")
        #版本5_2实现，提示输入后及时保存文件
        an3=input("新添加了信息，请及时保存！现在是否保存？(y/n)")
        if an3 in ("Y","y"):
            self.save_information()
            print("信息保存成功!")
        else:
            print("数据还未保存，请及时保存！")

    # 功能1：版本5_1,输出当前输入的多个工程师信息
    def output_information(self):
        print("-----------------------------------")
        print("输入的工程师的信息如下：")
        #遍历list2中存储的所有工程师信息，进行输出
        for infor in self.list2:
            print("编号:", infor["engineer_id"])
            print("姓名:", infor["engineer_name"])
            print("性别:", infor["engineer_sex"])
            print("工龄:", infor["engineer_work_year"])
            print("基本工资:", infor["engineer_base_salary"])
            print("绩效:", infor["engineer_performance"])
            print("月有效工作天数:", infor["engineer_workday"])
            print("保险费:", infor["engineer_InsuranceFee"])
            print("-----------------------------------")

    #功能2：版本5_1,删除指定工程师信息
    def del_information(self):
        an4="y"
        while an4 in("Y","y"):
            del_infor=input("请输入要删除的工程师编号或姓名:")
            #index=self.find_index(del_infor)
            #通过遍历，进行查找
            for infor in self.list2:
                #判断列表中是否存在指定删除的编号或姓名
                if del_infor in (infor["engineer_id"],infor["engineer_name"]):
                    confirm1=input("是否真的删除指定工程师信息？(y/n)")
                    if confirm1 in("Y","y"):
                        #得到指定删除的索引号
                        i=self.list2.index(infor)
                        #通过索引号进行删除指定信息
                        del self.list2[i]
                        print("数据删除成功！")
                    #删除后，后面不再判断
                    break
            else:
                print("没有找到指定的工程师编号或姓名！")
            an4=input("是否继续删除？(y/n)")

    # 功能3：版本6_1实现,查询工程师信息
    def find_information(self):
        an4 = "y"
        while an4 in ("Y", "y"):
            find_infor = input("请输入要查询的工程师编号或姓名:")
            for infor in self.list2:
                if find_infor in (infor["engineer_id"], infor["engineer_name"]):
                    i = self.list2.index(infor)
                    print("编号:", self.list2[i]["engineer_id"])
                    print("姓名:", self.list2[i]["engineer_name"])
                    print("性别:", self.list2[i]["engineer_sex"])
                    print("工龄:", self.list2[i]["engineer_work_year"])
                    print("基本工资:", self.list2[i]["engineer_base_salary"])
                    print("绩效:", self.list2[i]["engineer_performance"])
                    print("月有效工作天数:", self.list2[i]["engineer_workday"])
                    print("保险费:", self.list2[i]["engineer_InsuranceFee"])
                    print("每日餐补:", self.list2[i]["engineer_foodfee"])
                    print("月效益:", self.list2[i]["engineer_befitmonth"])
                    print("本月薪资:", self.list2[i]["total_salary"])
                    break
            else:
                print("没有找到指定的工程师编号或姓名！")
            an4 = input("是否继续查询？(y/n)")

    #功能4：版本5_1,修改指定工程师所有信息
    def update_information(self):
        an5 = "y"
        while an5 in ("Y", "y"):
            update_infor = input("请输入要修改的工程师编号或姓名:")
            # 通过遍历，进行查找
            for infor in self.list2:
                # 判断列表中是否存在指定修改的编号或姓名
                if update_infor in (infor["engineer_id"], infor["engineer_name"]):
                    # 得到指定修改的索引号
                    i = self.list2.index(infor)
                    print("1.修改全部信息\n2.修改部分信息")
                    select1 = input("请选择：")
                    if select1 == "1":
                        #调用修改全部函数
                        self.update_all_information(i)
                    elif select1 == "2":
                        #调用修改部分函数
                        self.update_partial_information(i)
                    else:
                        print("选择错误!")
                    # 修改后，后面不再判断
                    break
            else:
                print("没有找到指定的工程师编号或姓名！")
            #an5 = input("是否继续修改？(y/n)")

    #功能4，版本5_1,修改指定工程师的所有信息
    def update_all_information(self,i):
        print("-------------------------------")
        print("请修改软件测试工程师的信息:")
        #直接在list2中修改全部信息
        self.list2[i]["engineer_id"]= input("请输入新的编号:")
        self.list2[i]["engineer_name"] = input("请输入新的姓名：")
        self.list2[i]["engineer_sex"] = input("请输入新的性别：")
        self.list2[i]["engineer_work_year"] = input("请输入新的工龄：")
        self.list2[i]["engineer_base_salary"] = input("请输入新的基本工资：")
        self.list2[i]["engineer_performance"] = input("请输入新的绩效：")
        self.list2[i]["engineer_workday"] = input("请输入新的月有效工作天数：")
        self.list2[i]["engineer_InsuranceFee"] = input("请输入新的月保险费：")
        #调用修改餐补、月效益、薪资函数
        self.update_other(i)
        print("修改完毕！")

    #功能4：版本5_1,更新指定工程师部分信息
    def update_partial_information(self,i):
        print("请选择修改的内容：\n1.编号\n2.姓名\n3.性别\n4.工龄\n5.基本工资\n6.绩效\n7.月有效工作天数\n8.月保险费")
        an = "y"
        while an == "y" or an == "Y":
            update = input("请输入修改项：")
            if update == "1":
                self.list2[i]["engineer_id"] = input("请输入新的编号:")
            elif update == "2":
                self.list2[i]["engineer_name"] = input("请输入新的姓名：")
            elif update == "3":
                self.list2[i]["engineer_sex"] = input("请输入新的性别：")
            elif update == "4":
                self.list2[i]["engineer_work_year"] = input("请输入新的工龄：")
                #影响薪资的因素修改后，调用初始化函数
                self.update_other(i)
            elif update == "5":
                self.list2[i]["engineer_base_salary"] = input("请输入新的基本工资：")
                self.update_other(i)
            elif update == "6":
                self.list2[i]["engineer_performance"] = input("请输入新的绩效：")
                self.update_other(i)
            elif update == "7":
                self.list2[i]["engineer_workday"] = input("请输入新的月有效工作天数：")
                self.update_other(i)
            elif update == "8":
                self.list2[i]["engineer_InsuranceFee"] = input("请输入新的月保险费：")
                self.update_other(i)
            an=input("是否继续修改？(y/n)")
        else:
            print("修改完毕！")

    # 辅助函数，版本5_1实现，信息修改了，餐补，月效益和薪资要初始化
    def update_other(self,i):
        print("基本信息修改完毕，需要重新计算该工程师薪资!")
        self.list2[i]["engineer_foodfee"] = 0.0
        self.list2[i]["engineer_befitmonth"] = 0.0
        self.list2[i]["total_salary"] = 0.0

    # 功能5：版本5_1,计算指定工程师薪资，并输出
    def calculate_salary(self):
        print("计算薪资:")
        an3="y"
        while an3 in ("Y","y"):
            engineer=input("请输入工程师的编号或姓名:")
            #通过遍历，查找工程师信息
            for person in self.list2:
                #如果找到，进行计算
                if engineer==person["engineer_id"] or engineer==person["engineer_name"]:
                    i = self.list2.index(person)
                    self.list2[i]["engineer_foodfee"] = input("请输入每日餐补:")
                    self.list2[i]["engineer_befitmonth"] = input("请输入月效益：")
                    self.list2[i]["total_salary"] = (float(person["engineer_base_salary"]) + float(person["engineer_performance"]) +float(person["engineer_foodfee"]) * int(person["engineer_workday"]) +float(person["engineer_befitmonth"]) * int(person["engineer_work_year"]) / 100) * 0.9 - float(person["engineer_InsuranceFee"])
                    print("计算结果：")
                    print("---------------------------------------------------------------")
                    print("编号:%s\t|姓名:%s\t|本月薪资:%.2f" % (person["engineer_id"], person["engineer_name"],person["total_salary"]))
                    print("---------------------------------------------------------------")
                    break
            else:
                print("指定的编号或姓名不存在！")
            an3 = input("是否继续计算薪资？(y/n)")

    # 版本5_2实现，辅助函数，删除、修改、计算薪资后，提示是否保存
    def prompt_save(self):
        an = input("工程师信息已经更新，请及时保存！是否保存？(y/n)")
        if an in ("Y", "y"):
            self.save_information()
        else:
            print("信息更新后尚未保存！")

    #版本6.2实现，保存数据到excel文件
    def save_information(self):
        # 创建一个新的工作表,默认有一个表单Sheet
        wb = Workbook()
        #得到默认表单Sheet
        datasheet=wb.get_sheet_by_name("Sheet")
        #print(wb.get_sheet_names)
        #修改默认表单名称
        datasheet.title="软件测试工程师管理系统"
        #填写表头
        datasheet["A1"]="编号"
        datasheet["B1"] = "姓名"
        datasheet["C1"] = "性别"
        datasheet["D1"] = "工龄"
        datasheet["E1"] = "基本工资"
        datasheet["F1"] = "绩效"
        datasheet["G1"] = "工作天数"
        datasheet["H1"] = "保险费"
        datasheet["I1"] = "每日餐补"
        datasheet["J1"] = "月效益"
        datasheet["K1"] = "本月薪资"
        self.data=dict()
        for self.data in self.list2:#遍历出所有的工程师信息
            templist = list()
            for value in self.data.values():#遍历出每个工程师信息中的值
                templist.append(value)  #把每个工程师信息的值放到templist列表中
            #把列表中的值添加到excel的空行
            datasheet.append(templist)
        #保存文件
        wb.save("E:\mypythonpj\engineer.xlsx")

    #功能7：版本6_1实现，排序
    def sort_information(self):
        an7="y"
        while an7 in ("y","Y"):
            print("排序功能：(1)编号升序(2)姓名升序(3)工龄降序")
            select=input("请选择：")
            if select=="1":
                self.id_ascend()
            elif select=="2":
                self.name_ascend()
            elif select=="3":
                self.workyear_descend()
            else:
                print("选择错误！")
            an = input("是否保存排序后的信息？(y/n)")
            if an in ("Y", "y"):
                self.list2 = self.sort_list
                self.save_information()
            else:
                print("排序后信息尚未保存！")
            an7=input("是否继续排序？(y/n)")

    #编号升序
    def id_ascend(self):
        id_list=list()
        self.sort_list=list()
        num=len(self.list2)
        i=0
        while i<num:
            #把list2中的每个字典元素的编号取出来，放在id_list中
            id_list.append(int(self.list2[i]["engineer_id"]))
            i+=1
        #按照编号进行升序
        id_list.sort()
        for id in id_list:#遍历所有排好序的编号
            for item in self.list2:#遍历list2
                if id == int(item["engineer_id"]):
                    #排序好放在sort_list中
                    self.sort_list.append(item)
        #输出排序好的信息
        self.output_sort_information()

    #姓名升序
    def name_ascend(self):
        name_list=list()
        self.sort_list=list()
        num=len(self.list2)
        i=0
        while i<num:
            #把list2中的每个字典元素的姓名取出来，放在name_list中
            name_list.append(self.list2[i]["engineer_name"])
            i+=1
        #按照姓名进行升序
        name_list.sort()
        for name in name_list:#遍历所有排好序的姓名
            for item in self.list2:#遍历list2
                if name == item["engineer_name"]:
                    #排序好放在sort_list中
                    self.sort_list.append(item)
        #输出排序好的信息
        self.output_sort_information()

    #工龄升序
    def workyear_descend(self):
        workyear_list=list()
        self.sort_list=list()
        num=len(self.list2)
        i=0
        while i<num:
            #把list2中的每个字典元素的工龄取出来，放在workyear_list中
            workyear_list.append(int(self.list2[i]["engineer_work_year"]))
            i+=1
        #按照编号进行降序
        workyear_list.sort(reverse=True)
        for workyear in workyear_list:#遍历所有排好序的工龄
            for item in self.list2:#遍历list2
                if workyear == int(item["engineer_work_year"]):
                    #排序好放在sort_list中
                    self.sort_list.append(item)
        #输出排序好的信息
        self.output_sort_information()

    #辅助函数：排序后输出
    def output_sort_information(self):
        i=len(self.sort_list)
        if i !=0:
            print("按照编号升续输出信息如下：")
            for infor in self.sort_list:
                print("-----------------------------------")
                print("编号:", infor["engineer_id"])
                print("姓名:", infor["engineer_name"])
                print("性别:", infor["engineer_sex"])
                print("工龄:", infor["engineer_work_year"])
                print("基本工资:", infor["engineer_base_salary"])
                print("绩效:", infor["engineer_performance"])
                print("月有效工作天数:", infor["engineer_workday"])
                print("保险费:", infor["engineer_InsuranceFee"])
                print("每日餐补:", infor["engineer_foodfee"])
                print("月效益:", infor["engineer_befitmonth"])
                print("本月薪资:", infor["total_salary"])
        else:
            print("没有可输出的信息！")

    # 功能8：版本5_1实现，输出所有工程师信息
    def output_all_information(self):
        i=len(self.list2)
        if i !=0:
            print("已经存在的工程师信息如下：")
            for infor in self.list2:
                print("-----------------------------------")
                print("编号:", infor["engineer_id"])
                print("姓名:", infor["engineer_name"])
                print("性别:", infor["engineer_sex"])
                print("工龄:", infor["engineer_work_year"])
                print("基本工资:", infor["engineer_base_salary"])
                print("绩效:", infor["engineer_performance"])
                print("月有效工作天数:", infor["engineer_workday"])
                print("保险费:", infor["engineer_InsuranceFee"])
                print("每日餐补:", infor["engineer_foodfee"])
                print("月效益:", infor["engineer_befitmonth"])
                print("本月薪资:", infor["total_salary"])
        else:
            print("没有可输出的信息！")

    #功能9：清空工程师信息
    def clear_information(self):
        self.list2.clear()
        print("list2:",self.list2)
        print("数据清空完毕！")

    # 功能10：版本6_1实现，表格形式输出所有工程师信息
    def grid_output(self):
        i=len(self.list2)
        if i !=0:
            print("                                                                  软件测试工程师信息表")
            print("-------------------------------------------------------------------------------------------------------------------------------------------------")
            print("|\t编号\t|\t姓名\t|\t性别\t|\t工龄\t|\t基本工资\t|\t绩效\t|\t工作天数\t|\t保险\t|\t餐补\t|\t月效益\t|\t本月薪资\t|")
            print("-------------------------------------------------------------------------------------------------------------------------------------------------")

            for infor in self.list2:
                print("|\t",infor["engineer_id"],"\t|\t",infor["engineer_name"],"\t|\t",infor["engineer_sex"],"\t|\t",infor["engineer_work_year"],"\t|\t",infor["engineer_base_salary"],"\t\t|\t",infor["engineer_performance"],"\t|\t",infor["engineer_workday"],"\t\t|\t",infor["engineer_InsuranceFee"],"\t|\t",infor["engineer_foodfee"],"\t|\t",infor["engineer_befitmonth"],"\t|\t",infor["total_salary"],"\t|")
                print(  "-------------------------------------------------------------------------------------------------------------------------------------------------")
        else:
            print("没有可输出的信息！")
            
    #版本6.2实现，功能11，重新从excel文件导入数据
    def import_data(self):
        #打开指定的文件
        wb=load_workbook("E:\mypythonpj\engineer.xlsx")
        #通过sheet名称获得表单
        datasheet=wb.get_sheet_by_name("软件测试工程师管理系统")
        #存储所有的键
        listkey=("engineer_id","engineer_name","engineer_sex","engineer_work_year","engineer_base_salary","engineer_performance","engineer_workday","engineer_InsuranceFee","engineer_foodfee","engineer_befitmonth","total_salary")
        listvalue=list()
        for row in datasheet.rows:#遍历所有的行
            listvalue = list()
            for cell in row: #遍历每一行的每个单元格
                listvalue.append(cell.value) #把每一行的值存在一个列表中
            for key,value in zip(listkey,listvalue):
                self.engineer_information[key]=value #把键值存在字典中
            print(self.engineer_information)
            if self.engineer_information["engineer_id"]!="编号":#判断不是第一行
                self.list1 = copy.deepcopy(self.engineer_information)
                # 将列表1插入到列表2指定位置
                self.list2.append(self.list1)

    #初始化、清空数据
    def init_data(self):
        self.engineer_information["engineer_id"] = "1"
        self.engineer_information["engineer_name"] = ""
        self.engineer_information["engineer_sex"] = "男"
        self.engineer_information["engineer_work_year"] = 1
        self.engineer_information["engineer_base_salary"] = 0.0
        self.engineer_information["engineer_performance"] = 0.0
        self.engineer_information["engineer_workday"] = 0
        self.engineer_information["engineer_InsuranceFee"] = 0.0
        self.engineer_information["engineer_foodfee"] = 0.0
        self.engineer_information["engineer_befitmonth"] = 0.0
        self.engineer_information["total_salary"] = 0.0
